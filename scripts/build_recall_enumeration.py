#!/usr/bin/env python3
"""Build the human-anchored recall enumeration instrument.

Recall needs a human count of all true records on a page sample. This writes a sheet that,
for the SAME pages used in the automated estimate (same seed), lists the records the pipeline
extracted for each page and asks the reviewer to (a) mark any listed row that is NOT actually
on that page (spurious / mis-attributed) and (b) count records on the page the pipeline MISSED.

Recall is then: confirmed-present / (confirmed-present + missed), summed over pages -- a real
human number to replace the automated estimate.

Usage:  python scripts/build_recall_enumeration.py [--pages 25] [--seed 20260623]
"""
from __future__ import annotations

import argparse
import csv
import glob
import io
import json
import os
import random
import sys
import zipfile
from collections import defaultdict
from pathlib import Path

import openpyxl

sys.path.insert(0, str(Path(__file__).resolve().parents[1] / "src"))
from dark_health_data.config import settings  # noqa: E402


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--pages", type=int, default=25)
    ap.add_argument("--seed", type=int, default=20260623)
    ap.add_argument("--out", default="private/review-packet/eqr_recall_enumeration_basu.xlsx")
    args = ap.parse_args()

    z = sorted(glob.glob("dist/eqr-v*.zip"))[-1]
    zf = zipfile.ZipFile(z)
    recs = [json.loads(ln) for ln in zf.read([n for n in zf.namelist()
            if n.endswith("records.jsonl")][0]).decode().splitlines() if ln.strip()]
    docmeta = {d["document_id"]: d for d in csv.DictReader(io.StringIO(
        zf.read([n for n in zf.namelist() if n.endswith("documents.csv")][0]).decode()))}
    by_pp = defaultdict(list)
    for r in recs:
        if r.get("record_type") == "eqr_quality_measure":
            pv = r.get("provenance", {})
            if pv.get("page_start"):
                by_pp[(pv["source_document_id"], int(pv["page_start"]))].append(r)

    def locate(did):
        h = [x for x in glob.glob(os.path.join(str(settings.raw_dir), f"{did}*")) if not x.endswith(".meta.json")]
        return h[0] if h else None

    # SAME sampling as recall_estimate.py (seed + <=3 pages/doc) so human and automated align
    pages = [pp for pp in by_pp if len(by_pp[pp]) >= 1 and locate(pp[0])]
    rng = random.Random(args.seed)
    rng.shuffle(pages)
    by_doc, sample = defaultdict(int), []
    for pp in pages:
        if by_doc[pp[0]] < 3:
            sample.append(pp)
            by_doc[pp[0]] += 1
        if len(sample) >= args.pages:
            break

    wb = openpyxl.Workbook()
    ins = wb.active
    ins.title = "Instructions"
    for line in [
        "EQR recall enumeration — reviewer instructions",
        "",
        f"{len(sample)} pages. For each page, open prov_source_url to the listed page and look at",
        "every performance-measure value on that page (all rows of all tables; values in figures",
        "count too).",
        "",
        "The 'Records' tab lists what the pipeline extracted for each page, grouped by page.",
        "1. For each listed row, put 0 in `present_on_page` if that value is NOT actually on that",
        "   page (wrong page / spurious); leave blank if it is correctly on the page.",
        "2. On the 'PerPage' tab, for each page enter `n_missed` = how many performance-measure",
        "   values are ON the page but NOT in the pipeline's list for that page.",
        "",
        "Recall = (rows present) / (rows present + n_missed), summed across pages.",
    ]:
        ins.append([line])

    rec_ws = wb.create_sheet("Records")
    rec_ws.append(["page_id", "state", "page", "plan_name", "measure_name", "population",
                   "reporting_year", "rate", "present_on_page", "prov_source_url"])
    perpage = wb.create_sheet("PerPage")
    perpage.append(["page_id", "state", "page", "n_pipeline_records", "n_missed", "prov_source_url"])
    for i, (did, page) in enumerate(sample, 1):
        pid = f"pg-{i:02d}"
        url = docmeta.get(did, {}).get("source_url")
        st = docmeta.get(did, {}).get("jurisdiction")
        rows = by_pp[(did, page)]
        for r in rows:
            rec_ws.append([pid, st, page, r.get("plan_name"), r.get("measure_name"),
                           r.get("population"), r.get("reporting_year"), r.get("rate"), "", url])
        perpage.append([pid, st, page, len(rows), "", url])

    Path(args.out).parent.mkdir(parents=True, exist_ok=True)
    wb.save(args.out)
    print(f"wrote {len(sample)} pages, {sum(len(by_pp[pp]) for pp in sample)} pipeline records -> {args.out}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
