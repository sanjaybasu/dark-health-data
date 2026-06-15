#!/usr/bin/env python3
"""Targeted vision re-extraction of the rows a human reviewer scored INCORRECT.

Reads a completed EQR validation sheet (the reviewer's `correct` column), takes the
rows marked 0, and for each unique (source PDF, page) renders that page and runs the
Claude vision extractor over it. Emits a side-by-side table -- reviewer identity, the
text-pipeline value that was wrong, and the vision-read candidate value(s) for the same
measure -- so the reviewer can re-adjudicate whether vision recovers the figure.

Deliberately targeted: only the incorrect rows' pages are rendered (a few page-images),
so re-validation costs cents. Downloads are cached in /tmp by URL hash.

Usage:
  ANTHROPIC_API_KEY=... python scripts/vision_revalidate.py \
      "<completed_validation>.xlsx" [--model claude-haiku-4-5] [--out <path.csv>]
"""
from __future__ import annotations

import argparse
import csv
import hashlib
import sys
from pathlib import Path

import openpyxl

sys.path.insert(0, str(Path(__file__).resolve().parents[1] / "src"))

from dark_health_data.connectors.eqr import EQRConnector  # noqa: E402
from dark_health_data.extract.vision import ClaudeVisionExtractor  # noqa: E402
from dark_health_data.models import EQRQualityMeasure, SourceDocument  # noqa: E402

CACHE = Path("/tmp/dhd_revalidate")


def _download(url: str) -> Path | None:
    CACHE.mkdir(exist_ok=True)
    dest = CACHE / (hashlib.sha1(url.encode()).hexdigest()[:16] + ".pdf")
    if dest.exists() and dest.stat().st_size > 1000:
        return dest
    try:
        from dark_health_data.fetch import _download as fetch_bytes  # legacy-TLS aware
        dest.write_bytes(fetch_bytes(url))
        return dest
    except Exception as exc:
        print(f"  ! download failed ({exc}) for {url[:70]}", file=sys.stderr)
        return None


def _norm_measure(s: str) -> str:
    return "".join(c for c in (s or "").lower() if c.isalnum())


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("sheet")
    ap.add_argument("--model", default="claude-haiku-4-5")
    ap.add_argument("--out", default="private/review-packet/eqr_vision_readjudication.csv")
    args = ap.parse_args()

    wb = openpyxl.load_workbook(args.sheet, data_only=True)
    ws = wb["Review"]
    rows = list(ws.iter_rows(values_only=True))
    H = {h: i for i, h in enumerate([str(c) for c in rows[0]])}
    body = [r for r in rows[1:] if any(c is not None for c in r)]
    incorrect = [r for r in body if r[H["correct"]] == 0]
    print(f"{len(incorrect)} incorrect rows across "
          f"{len({(r[H['prov_source_url']], r[H['prov_page_start']]) for r in incorrect})} "
          f"unique (pdf, page) targets")

    connector = EQRConnector()
    extractor = ClaudeVisionExtractor(model=args.model)

    # group incorrect rows by (url, page) so each page is rendered once
    by_target: dict[tuple, list] = {}
    for r in incorrect:
        by_target.setdefault((r[H["prov_source_url"]], int(r[H["prov_page_start"]])), []).append(r)

    page_cache: dict[tuple, list] = {}  # (url,page) -> vision measures
    out_rows = []
    for (url, page), grp in by_target.items():
        if (url, page) not in page_cache:
            pdf_path = _download(url)
            if not pdf_path:
                page_cache[(url, page)] = []
            else:
                doc = SourceDocument(document_id=f"rv-{hashlib.sha1(url.encode()).hexdigest()[:8]}",
                                     dataset_id="eqr", source_url=url, local_path=str(pdf_path),
                                     jurisdiction=grp[0][H["state"]])
                extractor.pages = [page]
                recs = [r for r in extractor.extract("", doc, connector)
                        if isinstance(r, EQRQualityMeasure)]
                page_cache[(url, page)] = recs
                print(f"  {grp[0][H['state']]} p{page}: vision read {len(recs)} measure(s)")
        vis = page_cache[(url, page)]
        for r in grp:
            gm = _norm_measure(str(r[H["measure_name"]]))
            try:
                gy = int(r[H["reporting_year"]])
            except (TypeError, ValueError):
                gy = None
            # candidate vision values whose measure name overlaps the gold measure (either
            # direction, or sharing a >=6-char run). Model measure-naming drifts run-to-run,
            # so fall back to showing every value vision read on the page.
            def _overlaps(a: str, b: str) -> bool:
                if not a or not b:
                    return False
                if a in b or b in a:
                    return True
                return any(a[i:i + 6] in b for i in range(len(a) - 5))
            cands = [v for v in vis if _overlaps(_norm_measure(v.measure_name), gm)] or vis
            yr_cands = [v for v in cands if gy and v.reporting_year == gy and v.rate is not None]
            # the single value vision read for this exact (measure, year)
            exact = "; ".join(str(v.rate) for v in yr_cands) if yr_cands else ""
            series = "; ".join(
                f"{v.measure_name} [{v.reporting_year}]={v.rate}" for v in cands[:8]) or "(no figure match)"
            out_rows.append({
                "state": r[H["state"]], "measure_name": r[H["measure_name"]],
                "population": r[H["population"]], "reporting_year": gy,
                "text_pipeline_rate": r[H["rate"]],
                "vision_rate_same_year": exact,
                "vision_full_series_read": series,
                "page": page, "source_url": url,
                "reviewer_rescore_1correct_0incorrect": "",
            })

    out = Path(args.out)
    out.parent.mkdir(parents=True, exist_ok=True)
    cols = list(out_rows[0].keys())
    with open(out, "w", newline="") as fh:
        w = csv.DictWriter(fh, fieldnames=cols)
        w.writeheader()
        w.writerows(out_rows)

    # the clean, honest artifact: vision's full structured read of each figure page,
    # one row per value, for the reviewer to adjudicate against the rendered image.
    figpath = out.parent / "eqr_vision_figure_reads.csv"
    with open(figpath, "w", newline="") as fh:
        fcols = ["state", "page", "measure_name", "population", "reporting_year",
                 "vision_rate", "vision_numerator", "vision_denominator", "source_url"]
        w = csv.DictWriter(fh, fieldnames=fcols)
        w.writeheader()
        for (url, page), recs in page_cache.items():
            st = next((r[H["state"]] for r in incorrect
                       if r[H["prov_source_url"]] == url and int(r[H["prov_page_start"]]) == page), "")
            for v in recs:
                w.writerow({"state": st, "page": page, "measure_name": v.measure_name,
                            "population": v.population, "reporting_year": v.reporting_year,
                            "vision_rate": v.rate, "vision_numerator": v.numerator,
                            "vision_denominator": v.denominator, "source_url": url})
    print(f"  + per-figure structured read -> {figpath}")

    # a reviewer-friendly xlsx alongside the csv
    xlsx = out.with_suffix(".xlsx")
    wb2 = openpyxl.Workbook()
    sh = wb2.active
    sh.title = "Re-adjudicate"
    sh.append(cols)
    for r in out_rows:
        sh.append([r[c] for c in cols])
    wb2.save(xlsx)
    print(f"\nwrote {len(out_rows)} re-adjudication rows -> {out}\n                         and -> {xlsx}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
